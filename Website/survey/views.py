import os
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from openpyxl.utils import get_column_letter
from django.core.mail import EmailMessage

@csrf_exempt
def survey(request):
    email = 'nicholas.jarrett10@gmail.com'

    if request.method == 'POST':
        # Retrieve data from the form
        full_name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        age = request.POST.get('number', '')
        clarity_rating = request.POST.get('dropdown', '')
        strongest_language = request.POST.get('strongest-lang', '')
        strengths = request.POST.getlist('strengths')
        improvements = request.POST.get('textbox', '')

        # Excel file path
        excel_file_path = 'survey_responses.xlsx'

        # Create or load the workbook
        if not os.path.isfile(excel_file_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Reviews"  # Set the sheet name to "Reviews"
            # Write header row
            header = ['Full Name', 'Email', 'Age', 'Clarity Rating', 'Strongest Language',
                      'Strengths', 'Improvements']
            sheet.append(header)
        else:
            workbook = openpyxl.load_workbook(excel_file_path)
            if 'Reviews' not in workbook.sheetnames:
                workbook.create_sheet(title='Reviews')  # Create a new sheet with the name "Reviews"
            sheet = workbook['Reviews']

        # Add survey data to the worksheet
        strengths_text = ', '.join(strengths)
        survey_data = [full_name, email, age, clarity_rating, strongest_language,
                       strengths_text,
                       improvements]
        sheet.append(survey_data)

        # Iterate through each column and fit column width.
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Save the workbook after updating the cells
        workbook.save(excel_file_path)

        subject = 'New GitHub Portfolio Review'
        message = 'View the attached file to see it!'
        from_email = 'nicholas.jarrett10@gmail.com'
        recipient_list = [email]

        email = EmailMessage(subject, message, from_email, recipient_list)
        email.attach_file(excel_file_path)
        email.send()

        return HttpResponse('Form submitted successfully and data saved to Excel.')

    return render(request, 'survey.html')
